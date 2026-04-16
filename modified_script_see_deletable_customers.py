import pandas as pd
import warnings
import logging
import os

warnings.simplefilter(action="ignore", category=FutureWarning)

# --------------------------------------------------
# Configuration
# --------------------------------------------------
INPUT_FILE = "customers.csv"
OUTPUT_FILE = "customer_review.xlsx"

COL_ID = "id"
COL_NAME = "customer_name"
COL_ROADNAME = "customer_roadname"
COL_CUSTOMER_NUMBER = "customer_number"
COL_HANDLE = "dinero_handle"

DEFAULT_HANDLE = "unik"

# --------------------------------------------------
# Logging
# --------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")
logger = logging.getLogger(__name__)

# --------------------------------------------------
# Load file
# --------------------------------------------------
ext = os.path.splitext(INPUT_FILE)[1].lower()

if ext == ".xlsx":
    df = pd.read_excel(INPUT_FILE, engine="openpyxl")
elif ext == ".csv":
    df = pd.read_csv(
        INPUT_FILE,
        sep=";",
        encoding="cp1252",
        engine="python",
        on_bad_lines="warn"
    )
else:
    raise ValueError("Unsupported file type")

logger.info(f"Loaded {len(df)} rows")

# --------------------------------------------------
# Normalize column headers
# --------------------------------------------------
df.columns = (
    df.columns
    .str.strip()
    .str.replace('"', '', regex=False)
    .str.lower()
)

logger.info(f"Columns detected: {df.columns.tolist()}")

# --------------------------------------------------
# Validate required columns
# --------------------------------------------------
required = {COL_ID, COL_NAME, COL_ROADNAME, COL_CUSTOMER_NUMBER, COL_HANDLE}
missing = required - set(df.columns)

if missing:
    raise ValueError(f"Missing columns: {missing}")

# --------------------------------------------------
# Type conversions
# --------------------------------------------------
df[COL_ID] = pd.to_numeric(df[COL_ID], errors="raise")
df[COL_CUSTOMER_NUMBER] = pd.to_numeric(df[COL_CUSTOMER_NUMBER], errors="raise")

# --------------------------------------------------
# Normalize grouping keys
# --------------------------------------------------
df["_norm_name"] = df[COL_NAME].astype(str).str.strip().str.lower()
df["_norm_roadname"] = df[COL_ROADNAME].astype(str).str.strip().str.lower()

# --------------------------------------------------
# Process duplicate groups
# --------------------------------------------------
def process_group(group: pd.DataFrame) -> pd.DataFrame:
    group = group.sort_values(COL_ID)

    lowest_id = group.iloc[0][COL_ID]
    canonical_customer_number = group.iloc[0][COL_CUSTOMER_NUMBER]

    non_null_handles = group[COL_HANDLE].dropna()
    if not non_null_handles.empty:
        canonical_handle = non_null_handles.iloc[0]
    else:
        canonical_handle = DEFAULT_HANDLE

    group["canonical_customer_id"] = lowest_id
    group["canonical_customer_number"] = canonical_customer_number

    # Keep original dinero_handle untouched
    group["final_handle"] = None
    group.loc[group[COL_ID] == lowest_id, "final_handle"] = canonical_handle

    return group

df = (
    df.groupby(["_norm_name", "_norm_roadname"], group_keys=False)
      .apply(process_group)
      .reset_index(drop=True)
)

# --------------------------------------------------
# Mark duplicates
# --------------------------------------------------
df["to_delete"] = df[COL_ID] != df["canonical_customer_id"]

# --------------------------------------------------
# Sort (canonical first)
# --------------------------------------------------
df = df.sort_values(
    ["canonical_customer_id", "to_delete", COL_ID],
    ascending=[True, True, True]
).reset_index(drop=True)

# --------------------------------------------------
# Export to Excel with red highlighting
# --------------------------------------------------
from openpyxl.styles import PatternFill

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="customers")

    workbook = writer.book
    worksheet = writer.sheets["customers"]

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Find column index for "to_delete"
    col_index = df.columns.get_loc("to_delete") + 1

    for row_idx in range(2, len(df) + 2):  # Skip header
        if worksheet.cell(row=row_idx, column=col_index).value:
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_idx, column=col).fill = red_fill

logger.info(f"Excel file written: {OUTPUT_FILE}")
logger.info(f"Rows: {len(df)}")